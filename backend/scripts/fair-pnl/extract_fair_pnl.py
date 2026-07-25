import openpyxl, json, re, sys

FILES = [
    (r"C:\Users\User\Downloads\FAIR PNL Y'2025.xlsx", 2025),
    (r"C:\Users\User\Downloads\FAIR PNL Y'2026 (1).xlsx", 2026),
]

def norm(v):
    return str(v).strip() if v is not None else ""

def numify(v):
    if v is None: return None
    s = str(v).strip()
    if s == "" or s.startswith("#"): return None
    try: return round(float(s), 2)
    except: return None

def find_header(rows):
    """Return (idx, colmap) for the per-event table header row (has DATE + SALES + SET UP)."""
    for i, row in enumerate(rows):
        cells = [norm(c).upper() for c in row]
        if "DATE" in cells and "SALES" in cells and any("SET UP" in c or "SETUP" in c for c in cells):
            cm = {}
            for j, c in enumerate(cells):
                if c == "DATE" and "date" not in cm: cm["date"] = j
                elif c == "BRAND" and "brand" not in cm: cm["brand"] = j
                elif c == "LOCATION" and "location" not in cm: cm["location"] = j
                elif c == "SALES" and "sales" not in cm: cm["sales"] = j
                elif c == "RENTAL" and "rental" not in cm: cm["rental"] = j
                elif ("SET UP" in c or c == "SETUP") and "setup" not in cm: cm["setup"] = j
                elif "COMMISION" in c or "COMMISSION" in c: cm.setdefault("commission", j)
                elif "MERCHANT" in c: cm.setdefault("merchant", j)
            # COGS split lives in the 3 columns right after SALES (sub-header row below)
            if "sales" in cm:
                cm["cogs_matt_sofa"] = cm["sales"] + 1
                cm["cogs_bedframe"] = cm["sales"] + 2
                cm["cogs_accessories"] = cm["sales"] + 3
            return i, cm
    return None, None

events = []
for path, year in FILES:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        if "Sales" not in sn: continue
        ws = wb[sn]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        hidx, cm = find_header(rows)
        if hidx is None: continue
        # data starts 2 rows below header (skip the MATTRESS/SOFA sub-header row)
        for row in rows[hidx+2:]:
            date = norm(row[cm["date"]]) if cm["date"] < len(row) else ""
            brand = norm(row[cm["brand"]]) if cm["brand"] < len(row) else ""
            loc = norm(row[cm["location"]]) if cm["location"] < len(row) else ""
            sales = numify(row[cm["sales"]]) if cm["sales"] < len(row) else None
            if not date and not brand and not loc:  # blank separator
                continue
            if not brand or not loc:  # summary/footer noise
                continue
            def g(key):
                j = cm.get(key)
                return numify(row[j]) if j is not None and j < len(row) else None
            events.append({
                "sheet": sn, "year": year,
                "date_raw": date, "brand": brand, "location": loc,
                "sales": sales,
                "cogs_matt_sofa": g("cogs_matt_sofa"),
                "cogs_bedframe": g("cogs_bedframe"),
                "cogs_accessories": g("cogs_accessories"),
                "rental": g("rental"), "setup": g("setup"),
                "commission": g("commission"), "merchant": g("merchant"),
            })

out = r"C:\Users\User\AppData\Local\Temp\claude\C--Users-User-Desktop\074b956e-23d1-4528-8e33-82a9c7d1a85b\scratchpad\fair_pnl_events.json"
with open(out, "w", encoding="utf-8") as f:
    json.dump(events, f, ensure_ascii=False, indent=1)

# summary
withsales = [e for e in events if e["sales"]]
crossmonth = [e for e in events if re.search(r"/\d", e["date_raw"]) and "-" in e["date_raw"]]
brands = {}
for e in events: brands[e["brand"]] = brands.get(e["brand"], 0) + 1
print("total event rows:", len(events))
print("  with a sales figure:", len(withsales))
print("  cross-month date (has range):", len(crossmonth))
print("brands:", dict(sorted(brands.items(), key=lambda x:-x[1])))
print("sample 6:")
for e in events[:6]:
    print("  ", e["year"], e["date_raw"], "|", e["brand"], "|", e["location"][:28], "| sales", e["sales"], "| cogs(m/b/a)", e["cogs_matt_sofa"], e["cogs_bedframe"], e["cogs_accessories"], "| rent", e["rental"], "| setup", e["setup"])
print("wrote", out)
