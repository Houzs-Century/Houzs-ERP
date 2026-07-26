import openpyxl, json, re

FILES = [(r"C:\Users\User\Downloads\FAIR PNL Y'2025.xlsx", 2025),
         (r"C:\Users\User\Downloads\FAIR PNL Y'2026 (1).xlsx", 2026)]
MONTHS = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}
BRAND_FIX = {'DUNOPILLO':'DUNLOPILLO','AKEMI (C&C)':'AKEMI C&C'}

def norm(v): return str(v).strip() if v is not None else ""
def numify(v):
    if v is None: return 0.0
    s=str(v).strip()
    if s=="" or s.startswith("#"): return 0.0
    try: return float(s)
    except: return 0.0

def sheet_month(sn):
    m = re.search(r'([a-z]{3})', sn.lower())
    return MONTHS.get(m.group(1)) if m else None

def parse_dates(raw, endmonth, year):
    """'17-02/11' -> (start ISO, end ISO). SD>ED => start is previous month."""
    m = re.match(r'\s*(\d{1,2})\s*[-]\s*(\d{1,2})\s*/\s*(\d{1,2})', raw)
    if not m:
        m2 = re.match(r'\s*(\d{1,2})\s*/\s*(\d{1,2})', raw)  # single day 'DD/MM'
        if m2:
            d,mo=int(m2.group(1)),int(m2.group(2)); return (f"{year}-{mo:02d}-{d:02d}",)*2
        return (None,None)
    sd,ed,mo = int(m.group(1)),int(m.group(2)),int(m.group(3))
    emo=mo; smo=mo-1 if sd>ed else mo
    sy=year;
    if smo==0: smo=12; sy=year-1
    return (f"{sy}-{smo:02d}-{sd:02d}", f"{year}-{emo:02d}-{ed:02d}")

def find_header(rows):
    for i,row in enumerate(rows):
        cells=[norm(c).upper() for c in row]
        if "DATE" in cells and "SALES" in cells and any("SET UP" in c or c=="SETUP" for c in cells):
            cm={}
            for j,c in enumerate(cells):
                for key,lab in [("date","DATE"),("brand","BRAND"),("location","LOCATION"),("sales","SALES"),("rental","RENTAL")]:
                    if c==lab and key not in cm: cm[key]=j
                if ("SET UP" in c or c=="SETUP") and "setup" not in cm: cm["setup"]=j
            if "sales" in cm:
                cm["m"]=cm["sales"]+1; cm["b"]=cm["sales"]+2; cm["a"]=cm["sales"]+3
            return i,cm
    return None,None

raw=[]
for path,year in FILES:
    wb=openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        if "Sales" not in sn: continue
        emonth=sheet_month(sn)
        rows=[list(r) for r in wb[sn].iter_rows(values_only=True)]
        hi,cm=find_header(rows)
        if hi is None: continue
        for row in rows[hi+2:]:
            def cell(k):
                j=cm.get(k); return row[j] if j is not None and j<len(row) else None
            brand=norm(cell("brand")).upper()
            loc=norm(cell("location")); date=norm(cell("date"))
            sales_v=numify(cell("sales"))
            # END OF TABLE 1. The sheet repeats every event in a commission-claim
            # table below (its own DATE/BRAND header, preceded by a subtotal row).
            # Reading past table 1 double-counts every event -> stop here.
            if (date=="" and brand=="" and sales_v>0) or date.upper()=="DATE" or brand=="BRAND":
                break
            brand=BRAND_FIX.get(brand,brand)
            if not brand or not loc or loc.upper()=="LOCATION": continue
            org,venue=(loc.split("@",1)+[""])[:2] if "@" in loc else (loc,"")
            s,e=parse_dates(date, emonth, year)
            raw.append(dict(year=year,brand=brand,organizer=org.strip(),venue=venue.strip(),
                start=s,end=e,date_raw=date,
                sales=numify(cell("sales")),cogs_m=numify(cell("m")),cogs_b=numify(cell("b")),
                cogs_a=numify(cell("a")),rental=numify(cell("rental")),setup=numify(cell("setup"))))

# cross-month merge: same (brand, venue, organizer, start) collapses to one
merged={}
for e in raw:
    key=(e["brand"],e["venue"].upper(),e["organizer"].upper(),e["start"])
    if key in merged:
        for f in ["sales","cogs_m","cogs_b","cogs_a","rental","setup"]: merged[key][f]+=e[f]
        if e["end"] and (not merged[key]["end"] or e["end"]>merged[key]["end"]): merged[key]["end"]=e["end"]
    else: merged[key]=dict(e)
mg=list(merged.values())

def tot(rows,f): return sum(r[f] for r in rows)
print(f"raw rows: {len(raw)}   ->  merged projects: {len(mg)}")
for year in (2025,2026):
    ys=[r for r in mg if r["year"]==year]
    print(f"\n=== {year}  ({len(ys)} projects) ===")
    print(f"  SALES   RM {tot(ys,'sales'):,.0f}")
    print(f"  COGS    RM {tot(ys,'cogs_m')+tot(ys,'cogs_b')+tot(ys,'cogs_a'):,.0f}  (matt/sofa {tot(ys,'cogs_m'):,.0f} + bedframe {tot(ys,'cogs_b'):,.0f} + accessories {tot(ys,'cogs_a'):,.0f})")
    print(f"  RENTAL  RM {tot(ys,'rental'):,.0f}    SETUP RM {tot(ys,'setup'):,.0f}")
    bybrand={}
    for r in ys: bybrand[r["brand"]]=bybrand.get(r["brand"],0)+r["sales"]
    print("  sales by brand:", {k:f"{v:,.0f}" for k,v in sorted(bybrand.items(),key=lambda x:-x[1])})
