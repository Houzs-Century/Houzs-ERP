import openpyxl, json, re, collections

FILES=[(r"C:\Users\User\Downloads\FAIR PNL Y'2025.xlsx",2025),(r"C:\Users\User\Downloads\FAIR PNL Y'2026 (1).xlsx",2026)]
MONTHS={'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}
BRAND_FIX={'DUNOPILLO':'DUNLOPILLO','AKEMI (C&C)':'AKEMI C&C'}
# event type: Roadshow(=SOLO) vs Exhibition. Owner-confirmed 2026-07-26.
SOLO={'SOLO','VINCENT','SYELIN','MROOI','MALLMGMT','MALLMGT','KAIHAO','ROADSHOW','SUNWAYKLUANGMALLJOHOR'}
def sq(s): return re.sub(r'[^A-Z0-9]','',s.upper())
def etype(org):
    k=sq(org)
    return 'Roadshow' if ('SOLO' in k or k in SOLO) else 'Exhibition'
def norm(v): return str(v).strip() if v is not None else ""
def numf(v):
    if v is None: return 0.0
    s=str(v).strip()
    if s=="" or s.startswith("#"): return 0.0
    try: return round(float(s),2)
    except: return 0.0
def sheet_month(sn):
    m=re.search(r'([a-z]{3})',sn.lower()); return MONTHS.get(m.group(1)) if m else None
def parse_dates(raw,year):
    m=re.match(r'\s*(\d{1,2})\s*-\s*(\d{1,2})\s*/\s*(\d{1,2})',raw)
    if not m:
        m2=re.match(r'\s*(\d{1,2})\s*/\s*(\d{1,2})',raw)
        if m2: d,mo=int(m2.group(1)),int(m2.group(2)); return (f"{year}-{mo:02d}-{d:02d}",)*2
        return (None,None)
    sd,ed,mo=int(m.group(1)),int(m.group(2)),int(m.group(3))
    smo=mo-1 if sd>ed else mo; sy=year
    if smo==0: smo=12; sy=year-1
    return (f"{sy}-{smo:02d}-{sd:02d}",f"{year}-{mo:02d}-{ed:02d}")
def find_header(rows):
    for i,row in enumerate(rows):
        cells=[norm(c).upper() for c in row]
        if "DATE" in cells and "SALES" in cells and any("SET UP" in c or c=="SETUP" for c in cells):
            cm={}
            for j,c in enumerate(cells):
                for k,lab in [("date","DATE"),("brand","BRAND"),("location","LOCATION"),("sales","SALES"),("rental","RENTAL")]:
                    if c==lab and k not in cm: cm[k]=j
                if ("SET UP" in c or c=="SETUP") and "setup" not in cm: cm["setup"]=j
            if "sales" in cm: cm["m"],cm["b"],cm["a"]=cm["sales"]+1,cm["sales"]+2,cm["sales"]+3
            return i,cm
    return None,None

raw=[]
for path,year in FILES:
    wb=openpyxl.load_workbook(path,read_only=True,data_only=True)
    for sn in wb.sheetnames:
        if "Sales" not in sn: continue
        rows=[list(r) for r in wb[sn].iter_rows(values_only=True)]
        hi,cm=find_header(rows)
        if hi is None: continue
        for row in rows[hi+2:]:
            def cell(k):
                j=cm.get(k); return row[j] if j is not None and j<len(row) else None
            brand=norm(cell("brand")).upper(); loc=norm(cell("location")); date=norm(cell("date")); sv=numf(cell("sales"))
            if (date=="" and brand=="" and sv>0) or date.upper()=="DATE" or brand=="BRAND": break  # end of table 1
            brand=BRAND_FIX.get(brand,brand)
            if not brand or not loc or loc.upper()=="LOCATION": continue
            if "@" in loc:
                org,venue=(loc.split("@",1)+[""])[:2]
            elif loc.upper().startswith("SOLO"):
                org,venue="SOLO", loc[4:].strip(" -@")   # "SOLO AEON RAWANG" -> venue "AEON RAWANG"
            else:
                org,venue="", loc                          # no organizer; whole = venue
            s,e=parse_dates(date,year)
            raw.append(dict(year=year,brand=brand,organizer=org.strip(),venue=venue.strip(),event_type=etype(org),
                start=s,end=e,sales=sv,cogs_m=numf(cell("m")),cogs_b=numf(cell("b")),cogs_a=numf(cell("a")),
                rental=numf(cell("rental")),setup=numf(cell("setup"))))

# cross-month merge
mm={}
for e in raw:
    key=(e["brand"],e["venue"].upper(),e["organizer"].upper(),e["start"])
    if key in mm:
        for f in ["sales","cogs_m","cogs_b","cogs_a","rental","setup"]: mm[key][f]+=e[f]
        if e["end"] and (not mm[key]["end"] or e["end"]>mm[key]["end"]): mm[key]["end"]=e["end"]
    else: mm[key]=dict(e)
mg=list(mm.values())

# OWNER SPOT-CORRECTIONS to the source Excel (venue substring + start). The FAIR
# PNL sheet mis-recorded some events' brand. Add one entry per fix.
OVERRIDES=[
    {"venue_has":"PEX PAVILION BUKIT JALIL","start":"2026-03-06","set":{"brand":"MY SOFA FACTORY"}},
]
for r in mg:
    for ov in OVERRIDES:
        if ov["venue_has"] in r["venue"].upper() and r["start"]==ov["start"]:
            r.update(ov["set"])

# apportion setup + rental by sales across same-booth brands (venue,start,end)
# APPORTION only for ROADSHOW (SOLO): multiple brands share ONE booth -> split
# setup+rental by sales share. Exhibition brands have SEPARATE booths -> keep own.
booth=collections.defaultdict(list)
for r in mg: booth[(r["venue"].upper(),r["event_type"],r["start"],r["end"])].append(r)
for grp in booth.values():
    if grp[0]["event_type"]!="Roadshow": continue
    tot_sales=sum(r["sales"] for r in grp)
    if len(grp)<2 or tot_sales<=0: continue
    for f in ("setup","rental"):
        tot=sum(r[f] for r in grp)
        if tot<=0: continue
        for r in grp: r[f]=round(tot*r["sales"]/tot_sales,2)

# EXHIBITION reference-fill: an empty setup/rental borrows the AVERAGE from other
# Exhibition records at the SAME venue + SAME organizer (same fair) that have one.
ref=collections.defaultdict(lambda:{"setup":[],"rental":[]})
for r in mg:
    if r["event_type"]!="Exhibition": continue
    k=(r["venue"].upper(),r["organizer"].upper())
    if r["setup"]>0: ref[k]["setup"].append(r["setup"])
    if r["rental"]>0: ref[k]["rental"].append(r["rental"])
exfill=0
for r in mg:
    if r["event_type"]!="Exhibition": continue
    k=(r["venue"].upper(),r["organizer"].upper())
    for f in ("setup","rental"):
        if r[f]==0 and ref[k][f]:
            r[f]=round(sum(ref[k][f])/len(ref[k][f]),2); r[f"{f}_ref"]=True; exfill+=1
print(f"[exhibition reference-fill] filled {exfill} empty setup/rental from same venue+organizer")

# drop all-empty projects — but ONLY pre-Apr-2026. Owner 2026-07-26: Apr-2026
# onward is kept even if empty (data still being filled in; never delete those).
def anyamt(r): return r["sales"]+r["cogs_m"]+r["cogs_b"]+r["cogs_a"]+r["rental"]+r["setup"]
def keepable(r): return anyamt(r)>0 or (r["start"] or "9999")>="2026-04-01"
dropped=[r for r in mg if not keepable(r)]
mg=[r for r in mg if keepable(r)]
print(f"[drop] removed {len(dropped)} PRE-Apr-2026 all-empty projects (Apr-2026+ ALWAYS kept):")
for r in dropped: print(f"    {r['start']} {r['brand']} {r['organizer']} @ {r['venue'][:22]}")
print("[after apportion+drop, STILL missing]  revenue=0:", len([r for r in mg if not r['sales']]),
      " cogs=0:", len([r for r in mg if not (r['cogs_m']+r['cogs_b']+r['cogs_a'])]),
      " rental=0:", len([r for r in mg if not r['rental']]),
      " setup=0:", len([r for r in mg if not r['setup']]))

out=r"C:\Users\User\AppData\Local\Temp\claude\C--Users-User-Desktop\074b956e-23d1-4528-8e33-82a9c7d1a85b\scratchpad\seed_data_final.json"
json.dump(mg,open(out,"w",encoding="utf-8"),ensure_ascii=False,indent=1)

def tot(rows,f): return sum(r[f] for r in rows)
print(f"FINAL: {len(raw)} raw -> {len(mg)} projects")
print("event type:", dict(collections.Counter(r["event_type"] for r in mg)))
for y in (2025,2026):
    ys=[r for r in mg if r["year"]==y]
    print(f"\n{y}: {len(ys)} projects | SALES RM {tot(ys,'sales'):,.0f} | COGS RM {tot(ys,'cogs_m')+tot(ys,'cogs_b')+tot(ys,'cogs_a'):,.0f} | RENTAL RM {tot(ys,'rental'):,.0f} | SETUP RM {tot(ys,'setup'):,.0f}")
print(f"\napportion check -> projects with setup>0 now: {len([r for r in mg if r['setup']>0])} (was 371); rental>0: {len([r for r in mg if r['rental']>0])}")
print("wrote", out)
